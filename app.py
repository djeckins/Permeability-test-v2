from __future__ import annotations

import io
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from epidermal_barrier_screen.io import parse_input
from epidermal_barrier_screen.screen import screen_records

# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Permeability Screen",
    page_icon="🧪",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    [data-testid="stAppViewContainer"] {
        background: linear-gradient(168deg, #ffffff 0%, #f0f5fc 40%, #e4edf8 100%);
        font-family: 'Inter', sans-serif;
        color: #1a1a2e;
    }

    /* ── Global text readability ──────────────────── */
    p, span, label, div, .stMarkdown, .stText {
        color: #1a1a2e !important;
    }
    /* Force white text on dark (hero) background */
    .eb-hero, .eb-hero p, .eb-hero span, .eb-hero div, .eb-hero h1 {
        color: #ffffff !important;
    }
    .eb-hero p {
        color: #d0e4f5 !important;
    }
    /* Buttons: white text on blue */
    div[data-testid="stDownloadButton"] button,
    div[data-testid="stDownloadButton"] button span,
    div[data-testid="stButton"] button[kind="primary"],
    div[data-testid="stButton"] button[kind="primary"] span {
        color: #ffffff !important;
    }
    /* Metric boxes keep their theme colors */
    .m-total .val, .m-total .lbl         { color: #1565c0 !important; }
    .m-pass .val, .m-pass .lbl           { color: #2e7d32 !important; }
    .m-borderline .val, .m-borderline .lbl { color: #e65100 !important; }
    .m-fail .val, .m-fail .lbl           { color: #c62828 !important; }
    /* Criteria strip keeps its own colors */
    .eb-criteria-strip .strip-label      { color: #4a5a75 !important; }
    .eb-criteria-strip .eb-pill          { color: #1a3a65 !important; }
    [data-testid="stWidgetLabel"] label, [data-testid="stWidgetLabel"] p {
        color: #1a1a2e !important;
        font-weight: 500 !important;
    }
    .stTabs [data-baseweb="tab"] {
        color: #2a3a5c !important;
        font-weight: 600 !important;
    }
    .stTextArea textarea, .stNumberInput input {
        color: #1a1a2e !important;
    }
    .stTextArea textarea::placeholder {
        color: #7a8a9e !important;
    }
    [data-testid="stHeader"] { background: transparent; }

    /* Hide GitHub / Edit-source / Deploy toolbar buttons */
    [data-testid="stToolbar"]          { display: none !important; }
    [data-testid="stDecoration"]       { display: none !important; }
    #MainMenu                          { visibility: hidden !important; }
    footer                             { visibility: hidden !important; }

    /* ── Hero ─────────────────────────────────────── */
    .eb-hero {
        background: linear-gradient(135deg, #0f2b47 0%, #1a4578 40%, #2a6cb8 100%);
        border-radius: 20px;
        padding: 0;
        margin-bottom: 1.8rem;
        color: white;
        position: relative;
        overflow: hidden;
        min-height: 200px;
        box-shadow: 0 8px 32px rgba(15, 43, 71, 0.25), 0 2px 8px rgba(0,0,0,0.08);
    }
    .eb-hero-content {
        position: relative;
        z-index: 2;
        padding: 2.8rem 3rem 2.4rem;
        max-width: 60%;
    }
    .eb-hero h1 {
        font-size: 2.2rem;
        font-weight: 800;
        margin: 0 0 0.6rem;
        color: white;
        letter-spacing: -0.02em;
        line-height: 1.15;
    }
    .eb-hero p {
        font-size: 1.02rem;
        color: #b0cfe8;
        margin: 0;
        line-height: 1.55;
        font-weight: 400;
    }
    .eb-hero-art {
        position: absolute;
        top: 0;
        right: 0;
        width: 45%;
        height: 100%;
        z-index: 1;
        opacity: 0.35;
    }

    /* ── Criteria strip ───────────────────────────── */
    .eb-criteria-strip {
        background: white;
        border-radius: 16px;
        padding: 1.2rem 1.8rem;
        margin-bottom: 1.6rem;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        display: flex;
        align-items: center;
        gap: 0.7rem;
        flex-wrap: wrap;
    }
    .eb-criteria-strip .strip-label {
        font-size: 0.72rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        color: #4a5a75;
        margin-right: 0.4rem;
        white-space: nowrap;
    }
    .eb-pill {
        display: inline-flex;
        align-items: center;
        gap: 0.4rem;
        background: linear-gradient(135deg, #f0f5fc 0%, #e8eef8 100%);
        border: 1px solid #d5dff0;
        border-radius: 20px;
        padding: 0.38rem 0.85rem;
        font-size: 0.78rem;
        font-weight: 600;
        color: #1a3a65;
        transition: all 0.2s ease;
        white-space: nowrap;
    }
    .eb-pill:hover {
        background: linear-gradient(135deg, #e4edf8 0%, #d5dff0 100%);
        transform: translateY(-1px);
        box-shadow: 0 2px 8px rgba(42, 80, 128, 0.12);
    }
    .eb-pill .pill-icon {
        font-size: 0.82rem;
        opacity: 0.75;
    }

    /* ── Cards ────────────────────────────────────── */
    .eb-card {
        background: white;
        border-radius: 14px;
        padding: 1.5rem 1.8rem;
        box-shadow: 0 2px 12px rgba(0,0,0,0.06);
        margin-bottom: 1.4rem;
    }
    .eb-card h3 {
        margin: 0 0 1rem;
        font-size: 1rem;
        font-weight: 600;
        color: #1a3a5c;
    }

    /* ── Metric boxes ─────────────────────────────── */
    .eb-metric {
        border-radius: 14px;
        padding: 1.1rem 1.3rem;
        text-align: center;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }
    .eb-metric:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 16px rgba(0,0,0,0.10);
    }
    .eb-metric .val { font-size: 2.1rem; font-weight: 800; }
    .eb-metric .lbl {
        font-size: 0.72rem;
        text-transform: uppercase;
        letter-spacing: .07em;
        font-weight: 600;
        margin-top: 0.2rem;
    }
    .m-total      { background: linear-gradient(135deg, #e3f2fd, #d0e8fc); color: #1565c0; }
    .m-pass       { background: linear-gradient(135deg, #e8f5e9, #d0ecd2); color: #2e7d32; }
    .m-borderline { background: linear-gradient(135deg, #fff8e1, #fff0c0); color: #e65100; }
    .m-fail       { background: linear-gradient(135deg, #ffebee, #ffd6da); color: #c62828; }

    /* ── Buttons ──────────────────────────────────── */
    div[data-testid="stDownloadButton"] button {
        background: linear-gradient(135deg, #2563a8, #1a4578) !important;
        color: white !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 1rem !important;
        padding: .65rem 1.3rem !important;
        border: none !important;
        box-shadow: 0 3px 12px rgba(26, 69, 120, 0.3) !important;
        transition: all 0.2s ease !important;
    }
    div[data-testid="stDownloadButton"] button:hover {
        background: linear-gradient(135deg, #1a4578, #0f2b47) !important;
        box-shadow: 0 5px 18px rgba(26, 69, 120, 0.4) !important;
        transform: translateY(-1px) !important;
    }

    div[data-testid="stButton"] button[kind="primary"] {
        background: linear-gradient(135deg, #2563a8, #1a4578) !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        border: none !important;
        box-shadow: 0 3px 12px rgba(26, 69, 120, 0.25) !important;
        transition: all 0.2s ease !important;
    }
    div[data-testid="stButton"] button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1a4578, #0f2b47) !important;
        box-shadow: 0 5px 18px rgba(26, 69, 120, 0.35) !important;
        transform: translateY(-1px) !important;
    }

    /* ── Section headers ──────────────────────────── */
    h3 {
        color: #1a3a5c !important;
        font-weight: 700 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ── Constants ─────────────────────────────────────────────────────────────────
_STATUS_COLS = [
    "mw_status", "logd_status", "tpsa_status", "hbd_status",
    "hba_status", "rotb_status", "hac_status", "formal_charge_status",
    "ionization_status_criterion",
]

# Numeric column → its status column (for direct cell colouring)
_NUMERIC_STATUS = {
    "mw":                    "mw_status",
    "logd":                  "logd_status",
    "hbd":                   "hbd_status",
    "hba":                   "hba_status",
    "rotb":                  "rotb_status",
    "hac":                   "hac_status",
    "expected_net_charge":   "formal_charge_status",
    "fraction_unionized":    "ionization_status_criterion",
}

# Main visible columns in the Streamlit table
_DISPLAY_COLS = [
    "name",
    "mw",
    "hba",
    "hbd",
    "rotb",
    "hac",
    "predicted_pka",
    "clogp",
    "logd",
    "fraction_unionized",
    "final_result",
]

_CELL_CSS = {
    "optimal":    "background-color:#c8e6c9;color:#1b5e20",
    "suboptimal": "background-color:#fff9c4;color:#e65100",
    "poor":       "background-color:#ffcdd2;color:#b71c1c",
    "PASS":       "background-color:#a5d6a7;color:#1b5e20;font-weight:bold",
    "BORDERLINE": "background-color:#ffe082;color:#e65100;font-weight:bold",
    "FAIL":       "background-color:#ef9a9a;color:#b71c1c;font-weight:bold",
    "invalid_input": "background-color:#e0e0e0;color:#616161;font-weight:bold",
}

_XLSX_FILLS = {
    "optimal":    PatternFill("solid", fgColor="C8E6C9"),
    "suboptimal": PatternFill("solid", fgColor="FFF9C4"),
    "poor":       PatternFill("solid", fgColor="FFCDD2"),
    "PASS":       PatternFill("solid", fgColor="A5D6A7"),
    "BORDERLINE": PatternFill("solid", fgColor="FFE082"),
    "FAIL":       PatternFill("solid", fgColor="EF9A9A"),
    "invalid_input": PatternFill("solid", fgColor="E0E0E0"),
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _style_df(df: pd.DataFrame):
    def _color(val):
        return _CELL_CSS.get(str(val), "")

    styler = df.style

    # Colour final_result cell
    if "final_result" in df.columns:
        styler = styler.map(_color, subset=["final_result"])

    # Colour numeric cells directly based on their status column
    for num_col, status_col in _NUMERIC_STATUS.items():
        if num_col in df.columns and status_col in df.columns:
            styler = styler.apply(
                lambda col, sc=status_col: [
                    _CELL_CSS.get(str(df.at[i, sc]), "") for i in col.index
                ],
                subset=[num_col],
            )

    return styler


def _build_xlsx(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)
    wb = load_workbook(buf)
    ws = wb.active

    # Header styling
    hdr_fill = PatternFill("solid", fgColor="1A3A5C")
    hdr_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 36

    # Resolve which columns need colour
    headers = [c.value for c in ws[1]]
    colour_cols = {
        col: headers.index(col) + 1
        for col in _STATUS_COLS + ["final_result"]
        if col in headers
    }

    # Apply row colours
    for row in ws.iter_rows(min_row=2):
        for col_name, col_idx in colour_cols.items():
            cell = row[col_idx - 1]
            fill = _XLSX_FILLS.get(str(cell.value))
            if fill:
                cell.fill = fill
        # Zebra stripe on non-coloured cells
        for cell in row:
            if cell.column not in colour_cols.values():
                if cell.row % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F5F8FF")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 32)

    # Freeze header
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _detect_mode(filename: str) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".zip":
        return "sdf_zip"
    if suffix in (".sdf", ".mol"):
        return "sdf"
    return "smiles_list"   # .txt or anything else


# ── UI ────────────────────────────────────────────────────────────────────────

# Decorative SVG for hero section – abstract hexagons, molecular network, floating spheres, wave layers
_HERO_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 500 260" width="100%" height="100%" preserveAspectRatio="xMaxYMid slice">
  <defs>
    <linearGradient id="wg" x1="0" y1="0" x2="0" y2="1">
      <stop offset="0%" stop-color="#ffffff" stop-opacity="0.10"/>
      <stop offset="100%" stop-color="#ffffff" stop-opacity="0.02"/>
    </linearGradient>
    <radialGradient id="sg" cx="50%" cy="50%" r="50%">
      <stop offset="0%" stop-color="#ffffff" stop-opacity="0.25"/>
      <stop offset="100%" stop-color="#ffffff" stop-opacity="0"/>
    </radialGradient>
  </defs>

  <!-- Translucent wave layers -->
  <path d="M0,220 Q80,180 180,200 T360,185 T500,210 L500,260 L0,260Z" fill="url(#wg)" opacity="0.6"/>
  <path d="M0,235 Q120,205 240,225 T480,210 L500,260 L0,260Z" fill="url(#wg)" opacity="0.4"/>

  <!-- Molecular network lines -->
  <g stroke="#ffffff" stroke-opacity="0.12" stroke-width="1" fill="none">
    <line x1="180" y1="60" x2="260" y2="95"/>
    <line x1="260" y1="95" x2="350" y2="70"/>
    <line x1="350" y1="70" x2="420" y2="110"/>
    <line x1="260" y1="95" x2="290" y2="160"/>
    <line x1="290" y1="160" x2="380" y2="175"/>
    <line x1="380" y1="175" x2="420" y2="110"/>
    <line x1="180" y1="60" x2="150" y2="130"/>
    <line x1="150" y1="130" x2="220" y2="170"/>
    <line x1="220" y1="170" x2="290" y2="160"/>
    <line x1="350" y1="70" x2="440" y2="50"/>
    <line x1="440" y1="50" x2="480" y2="90"/>
    <line x1="420" y1="110" x2="480" y2="90"/>
  </g>

  <!-- Network node dots -->
  <g fill="#ffffff" fill-opacity="0.25">
    <circle cx="180" cy="60" r="3.5"/>
    <circle cx="260" cy="95" r="4"/>
    <circle cx="350" cy="70" r="3"/>
    <circle cx="420" cy="110" r="3.5"/>
    <circle cx="290" cy="160" r="3"/>
    <circle cx="380" cy="175" r="3.5"/>
    <circle cx="150" cy="130" r="2.5"/>
    <circle cx="220" cy="170" r="3"/>
    <circle cx="440" cy="50" r="2.5"/>
    <circle cx="480" cy="90" r="3"/>
  </g>

  <!-- Abstract hexagons -->
  <g fill="none" stroke="#ffffff" stroke-opacity="0.10" stroke-width="1.2">
    <polygon points="320,35 345,22 370,35 370,60 345,73 320,60"/>
    <polygon points="370,60 395,47 420,60 420,85 395,98 370,85"/>
    <polygon points="200,120 220,110 240,120 240,140 220,150 200,140"/>
    <polygon points="400,155 418,145 436,155 436,173 418,183 400,173"/>
  </g>

  <!-- Floating translucent spheres -->
  <circle cx="450" cy="45" r="18" fill="url(#sg)" opacity="0.5"/>
  <circle cx="160" cy="80" r="12" fill="url(#sg)" opacity="0.4"/>
  <circle cx="330" cy="190" r="22" fill="url(#sg)" opacity="0.35"/>
  <circle cx="470" cy="170" r="10" fill="url(#sg)" opacity="0.45"/>
  <circle cx="240" cy="40" r="8" fill="url(#sg)" opacity="0.3"/>
</svg>"""

st.markdown(
    f"""
    <div class="eb-hero">
      <div class="eb-hero-art">{_HERO_SVG}</div>
      <div class="eb-hero-content">
        <h1>Permeability Screening Tool</h1>
        <p>Evaluate compounds for membrane permeation potential.<br>
           Supports single SMILES, SMILES lists, SDF files, and ZIP archives.</p>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Criteria overview strip ──────────────────────────────────────────────────
_CRITERIA_PILLS = [
    ("⚖️", "MW"),
    ("🧪", "LogP / LogD"),
    ("◉", "TPSA"),
    ("🔗", "HBD & HBA"),
    ("↻", "Rotatable Bonds"),
    ("⬢", "Heavy Atom Count"),
    ("±", "Formal Charge"),
    ("⚡", "Ionization at pH"),
]
_pills_html = "".join(
    f'<span class="eb-pill"><span class="pill-icon">{icon}</span>{label}</span>'
    for icon, label in _CRITERIA_PILLS
)
st.markdown(
    f"""
    <div class="eb-criteria-strip">
      <span class="strip-label">Screening Criteria</span>
      {_pills_html}
    </div>
    """,
    unsafe_allow_html=True,
)

# ── Input section ─────────────────────────────────────────────────────────────
tab_smiles, tab_file = st.tabs(["✏️  Enter SMILES", "📂  Upload File"])

smiles_payload: str | None = None
file_payload: bytes | None = None
file_name: str | None = None

with tab_smiles:
    st.markdown(
        "Enter a **SMILES**, a **compound name**, or a **list** (one per line).\n"
        "Names are automatically resolved to structures via PubChem."
    )
    smiles_text = st.text_area(
        label="SMILES or name input",
        placeholder="caffeic acid\nCC(=O)Oc1ccccc1C(=O)O  Aspirin\nibuprofen\nCCO  Ethanol",
        height=160,
        label_visibility="collapsed",
    )
    if smiles_text and smiles_text.strip():
        smiles_payload = smiles_text.strip()

with tab_file:
    st.markdown("Upload an **SDF file**, a **ZIP archive** containing SDFs, or a **TXT file** with SMILES:")
    uploaded = st.file_uploader(
        label="File upload",
        type=["sdf", "zip", "txt"],
        label_visibility="collapsed",
    )
    if uploaded is not None:
        file_payload = uploaded.getvalue()
        file_name = uploaded.name

# ── pH input (required) ───────────────────────────────────────────────────────
_, ph_col, _ = st.columns([2, 3, 2])
with ph_col:
    ph_input = st.number_input(
        "pH for ionization assessment",
        min_value=0.0,
        max_value=14.0,
        value=None,
        step=0.1,
        format="%.1f",
        placeholder="Enter pH (0–14)",
    )

# ── Run button ────────────────────────────────────────────────────────────────
st.markdown("<div style='height:0.4rem'/>", unsafe_allow_html=True)
_, btn_col, _ = st.columns([2, 3, 2])
with btn_col:
    run = st.button("🔬  Run Screening", type="primary", use_container_width=True)

# ── Processing ────────────────────────────────────────────────────────────────
if run:
    if ph_input is None:
        st.error("Введите pH перед запуском расчёта.")
        st.stop()

    # File takes priority over SMILES text
    if file_payload is not None:
        mode = _detect_mode(file_name or "upload.sdf")
        payload = file_payload
        fname = file_name
    elif smiles_payload:
        mode = "smiles_list" if "\n" in smiles_payload else "smiles"
        payload = smiles_payload
        fname = None
    else:
        st.error("Please enter a SMILES string or upload a file before running.")
        st.stop()

    with st.spinner("Calculating descriptors and screening…"):
        records = parse_input(mode, payload, filename=fname)
        df = screen_records(records, ph=float(ph_input))

    # ── Metrics ───────────────────────────────────────────────────────────────
    total      = len(df)
    n_pass      = int((df["final_result"] == "PASS").sum())
    n_border    = int((df["final_result"] == "BORDERLINE").sum())
    n_fail      = int((df["final_result"] == "FAIL").sum())
    n_invalid   = int((df["parse_status"] != "ok").sum())

    st.markdown("<div style='height:0.4rem'/>", unsafe_allow_html=True)
    m1, m2, m3, m4 = st.columns(4)
    for col, cls, label, value in [
        (m1, "m-total",      "Total",      total),
        (m2, "m-pass",       "PASS",        n_pass),
        (m3, "m-borderline", "BORDERLINE",  n_border),
        (m4, "m-fail",       "FAIL",        n_fail),
    ]:
        col.markdown(
            f'<div class="eb-metric {cls}">'
            f'<div class="val">{value}</div>'
            f'<div class="lbl">{label}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    if n_invalid:
        st.warning(f"{n_invalid} record(s) could not be parsed and are shown as `invalid_input`.")

    # ── Results table ─────────────────────────────────────────────────────────
    st.markdown("<div style='height:0.8rem'/>", unsafe_allow_html=True)
    st.markdown("### Results")
    display_order = [c for c in _DISPLAY_COLS if c in df.columns]
    st.dataframe(
        _style_df(df),
        use_container_width=True,
        hide_index=True,
        height=min(40 + len(df) * 36, 600),
        column_order=display_order,
        column_config={
            "name": st.column_config.TextColumn("Compound"),
            "mw": st.column_config.NumberColumn("MW"),
            "hba": st.column_config.NumberColumn("HBA"),
            "hbd": st.column_config.NumberColumn("HBD"),
            "rotb": st.column_config.NumberColumn("RB"),
            "hac": st.column_config.NumberColumn("HAC"),
            "predicted_pka": st.column_config.NumberColumn("pKa"),
            "clogp": st.column_config.NumberColumn("cLogP"),
            "logd": st.column_config.NumberColumn(f"LogD (pH {ph_input:.1f})"),
            "fraction_unionized": st.column_config.NumberColumn(
                f"f_unionized (pH {ph_input:.1f})"
            ),
            "final_result": st.column_config.TextColumn("Final Result"),
        },
    )

    # ── Download ──────────────────────────────────────────────────────────────
    st.markdown("<div style='height:0.6rem'/>", unsafe_allow_html=True)
    xlsx_bytes = _build_xlsx(df)
    _, dl_col, _ = st.columns([2, 3, 2])
    with dl_col:
        st.download_button(
            label="⬇️  Download Results (Excel)",
            data=xlsx_bytes,
            file_name="permeability_screening_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ── Criteria reference ────────────────────────────────────────────────────────
st.markdown("<div style='height:1rem'/>", unsafe_allow_html=True)
with st.expander("📋  Screening criteria reference"):
    current_ph = ph_input if ph_input is not None else 5.5
    st.markdown(
        f"""
        Per-criterion classification: 🟢 **optimal** · 🟡 **suboptimal** · 🔴 **poor**

        Overall result:
        - **PASS** = at most **1 non-optimal** criterion
        - **BORDERLINE** = more than 1 non-optimal criterion and at most **2 poor** criteria
        - **FAIL** = **3 or more poor** criteria

        | Criterion | Optimal | Suboptimal | Poor |
        |:---|:---:|:---:|:---:|
        | **MW** | < 300 Da | 300–500 Da | > 500 Da |
        | **LogD / cLogP** | 1–3 | 0.5–1 or 3–5 | < 0.5 or > 5 |
        | **TPSA** | < 60 Å² | 60–130 Å² | > 130 Å² |
        | **HBD** | 0–3 | 4–5 | > 5 |
        | **HBA** | 2–8 | 8–10 | > 10 |
        | **RotB** | < 10 | 10–15 | > 15 |
        | **HAC** | < 30 | 30–50 | > 50 |
        | **Charge at pH {current_ph:.1f}** | neutral / near-neutral | about ±1 | multi-charged |
        | **Ionization (pH {current_ph:.1f})** | f_unionized ≥ 0.8 | 0.5–0.8 | < 0.5 |

        Name resolution: compound names are resolved to SMILES via **PubChem PUG REST**.

        pKa pipeline in this build:
        1. `input_pka_acidic` / `input_pka_basic` / `input_pka` from the file
        2. **PubChem** PUG-View lookup (Dissociation Constants annotations)
        3. **DrugBank** web lookup by compound name (public page parsing)
        4. **ChEMBL** API lookup (molecule properties, InChIKey / SMILES / name)
        5. **QupKake** ML-based site-aware pKa predictor (if installed, DB-fallback only)
        6. **Dimorphite-DL** protonation state enumeration at the selected pH
        7. pH-specific ionization and logD calculation

        The main pKa column shows a single value only when chemically meaningful
        (simple monoprotic acid/base). For multiprotic, ampholytic, or polyphenolic
        molecules the pKa column is left blank — detailed pKa data is in the export.

        `logD_7_4` from the input file is only used as an override when the selected pH is ~7.4.
        Otherwise the app reports a pH-corrected estimate derived from the selected pH.
        """
    )
